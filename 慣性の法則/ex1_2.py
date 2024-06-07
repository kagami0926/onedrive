import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib
import seaborn as sns
import numpy as np

# Excelファイルからデータを読み込む
wb = openpyxl.load_workbook("ex1.xlsx")
ws = wb["Sheet2"]
row_max = ws.max_row

# データの集合v1とv2を設定
v1 = []
v2 = []

# データを取得する
for row in ws.iter_rows(min_row=2, max_row=row_max, min_col=1, max_col=2, values_only=True):
    v1.append(row[0])  # 台車の速度(km/h)のデータを追加
    v2.append(row[1])  # 試走車の速度(km/h)のデータを追加

# 散布図を作成する
plt.figure(figsize=(10, 6))  # 図のサイズを指定する
p = sns.scatterplot(x=v1, y=v2, legend=False)  # 凡例を非表示にする
sns.set_style("whitegrid", {'grid.linestyle': '--'})

# ラベルを設定する(LaTeX形式)
p.set_xlabel("台車の速度"r"$V$""(km/h)", fontsize=16)
p.set_ylabel("試走車の速度"r"$v$""(km/h)", fontsize=16)
sns.set(font='IPAexGothic')

# 最小二乗法による直線のフィッティング
a = np.round(np.dot(v1, v2) / np.dot(v1, v1), 4)

# フィッティング直線の式
fit_line_x = np.linspace(0, max(v1), 100)  # 原点(0)から始める
fit_line_y = a * fit_line_x

# フィッティング直線をプロットする
plt.plot(fit_line_x, fit_line_y, color="green", label=f"近似直線"r"$(v=1.19V)$")

# 追加：v2=3.480v1 の直線をプロットする
theoretical_line_y = 3.480 * fit_line_x
plt.plot(fit_line_x, theoretical_line_y, color="blue", linestyle="--", label="理論値"r"$v=3.48V$")

# x軸とy軸の範囲を設定して線を長く表示する
plt.xlim(0, max(v1) + 1)
plt.ylim(0, max(v2) + 1)

# 凡例を表示
plt.legend(prop={'size': 15})

# 結果を表示する
plt.grid()
plt.show()

# 傾きを表示する
print("傾き:", a)
