import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib
import seaborn as sns
import numpy as np

# Excelファイルからデータを読み込む
wb = openpyxl.load_workbook("ex1.xlsx")
ws = wb["Sheet3"]
row_max = ws.max_row

# データの集合mとvを設定
m = []
v = []

# データを取得する
for row in ws.iter_rows(min_row=2, max_row=row_max, min_col=1, max_col=2, values_only=True):
    m.append(row[0])  # 錘の質量m(kg)のデータを追加
    v.append(row[1])  # 台車の速度V(km/h)のデータを追加

# 散布図を作成する
plt.figure(figsize=(10, 6))  # 図のサイズを指定する
p = sns.scatterplot(x=m, y=v, legend=False)  # 凡例を非表示にする
sns.set_style("whitegrid", {'grid.linestyle': '--'})

# ラベルを設定する(LaTeX形式)
p.set_xlabel("錘の質量"r"$m$""(kg)", fontsize=16)
p.set_ylabel("台車の速度の二乗"r"$V^2(km^2/h^2)$", fontsize=16)
sns.set(font='IPAexGothic')

# 最小二乗法による直線のフィッティング
a = np.dot(m, v) / np.dot(m, m)
a = round(a, 2)

# フィッティング直線の式
fit_line_x = np.linspace(0, max(m), 100)  # 原点(0)から始める
fit_line_y = a * fit_line_x

# フィッティング直線をプロットする
plt.plot(fit_line_x, fit_line_y, color="green", label=f"近似直線 "r"$(V^2=50.7m)$")

# 追加：V^2=72.63m の直線をプロットする
theoretical_line_y = 72.63 * fit_line_x
plt.plot(fit_line_x, theoretical_line_y, color="blue", linestyle="--", label="理論値"r"$V^2=72.6m$")

# x軸とy軸の範囲を設定して線を長く表示する
plt.xlim(0, max(m) + 0.05)
plt.ylim(0, max(v) + 0.05)

# 凡例を表示
plt.legend(prop={'size': 15})

# 結果を表示する
plt.grid()
plt.show()

# 傾きを表示する
print("傾き:", a)
