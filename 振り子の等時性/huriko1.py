import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib
import seaborn as sns
import numpy as np

# Excelファイルからデータを読み込む
wb = openpyxl.load_workbook("振り子の等時性\data\huriko.xlsx")

ws = wb["Sheet3"]
row_max = ws.max_row

# データの集合thとTを設定
th = []
T = []

# データを取得する
for row in ws.iter_rows(min_row=2, max_row=row_max, min_col=1, max_col=2, values_only=True):
    th.append(row[0])  # 振れ角のデータを追加
    T.append(row[1])  # 周期のデータを追加

# 散布図を作成する
plt.figure(figsize=(10, 6))  # 図のサイズを指定する
p = sns.scatterplot(x=th, y=T, legend=False)  # 凡例を非表示にする
sns.set_style("whitegrid", {'grid.linestyle': '--'})

# ラベルを設定する(LaTeX形式)
p.set_xlabel("振れ角"r"$\theta$""(度)", fontsize=16)
p.set_ylabel("周期"r"$T$""(s)", fontsize=16)
sns.set(font='IPAexGothic')

# 結果を表示する
plt.grid()
plt.show()

pl


