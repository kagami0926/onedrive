import openpyxl
import matplotlib.pyplot as plt
import japanize_matplotlib
import seaborn as sns
import numpy as np

# Excelファイルからデータを読み込む
wb = openpyxl.load_workbook("振り子の等時性/data/huriko.xlsx")
ws = wb["Sheet4"]
row_max = ws.max_row

# データの集合lとT2を設定
l = []
T2 = []

# データを取得する
for row in ws.iter_rows(min_row=2, max_row=row_max, min_col=1, max_col=2, values_only=True):
    l.append(row[0])  # 振れ角のデータを追加
    T2.append(row[1])  # 周期のデータを追加

# 散布図を作成する
plt.figure(figsize=(10, 6))  # 図のサイズを指定する
p = sns.scatterplot(x=l, y=T2, legend=False)  # 凡例を非表示にする
sns.set_style("whitegrid", {'grid.linestyle': '--'})

# ラベルを設定する(LaTeX形式)
p.set_xlabel("振れ角"r"$\theta$""(度)", fontsize=16)
p.set_ylabel("周期"r"$T^2$""(s^2)", fontsize=16)
sns.set(font='IPAexGothic')

# 原点を通る最小二乗法による直線のフィッティング
l = np.array(l)
T2 = np.array(T2)
a = np.round(np.dot(l, T2) / np.dot(l, l), 4)

# フィッティング直線の式
fit_line_x = np.linspace(0, max(l) * 1.5, 100)  # 原点(0)から最大値の1.5倍まで
fit_line_y = a * fit_line_x

# フィッティング直線をプロットする
plt.plot(fit_line_x, fit_line_y, color="green", label=f"$T^2 = {a}\\theta$")

# グラフのタイトルと凡例を追加
plt.title("振れ角と周期の関係", fontsize=18)
plt.legend(fontsize=14)  # 凡例のフォントサイズを設定

# x軸とy軸を原点から表示するように設定
plt.xlim(0, max(l) * 1.5)  # x軸の範囲を0から最大値の1.5倍まで
plt.ylim(0, max(T2) * 1.5)  # y軸の範囲を0から最大値の1.5倍まで

# グリッドを表示
plt.grid()

# グラフを画像ファイルとして保存
plt.savefig("振り子の等時性/images/huriko_graph.png")

# 結果を表示する
plt.show()
